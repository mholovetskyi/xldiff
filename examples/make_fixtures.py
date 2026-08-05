import os

from openpyxl import Workbook

QCOLS = ["C", "D", "E", "F"]


def build(version):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue build"

    ws["A1"] = "Revenue build"
    ws["B3"] = "Quarter"
    for i, c in enumerate(QCOLS):
        ws["%s3" % c] = "Q%d" % (i + 1)

    ws["B5"] = "Growth rate"
    ws["C5"] = 0.03 if version == 14 else 0.05

    ws["B9"] = "Base units"
    for c, v in zip(QCOLS, [1000, 1050, 1100, 1150]):
        ws["%s9" % c] = v

    ws["B10"] = "Units"
    for c in QCOLS:
        ws["%s10" % c] = "={c}9*(1+$C$5)".format(c=c)

    ws["B12"] = "Revenue"
    for c in QCOLS:
        ws["%s12" % c] = "={c}9*{c}22".format(c=c)
    if version == 14:
        ws["E12"] = "=E9*E23"

    ws["B13"] = "Churn cost"
    for c in QCOLS:
        ws["%s13" % c] = "={c}12*0.05".format(c=c)
    if version == 15:
        ws["F13"] = "=F12*0.18"

    ws["B14"] = "Total revenue"
    if version == 14:
        ws["C14"] = "=SUM(C12:F12)"
    else:
        ws["C14"] = 1483200

    ws["B22"] = "Price"
    for c, v in zip(QCOLS, [120, 120, 125, 125]):
        ws["%s22" % c] = v

    ws["B23"] = "Discount factor"
    for c in QCOLS:
        ws["%s23" % c] = 0.9

    op = wb.create_sheet("Opex")
    op["A1"] = "Operating expenses"
    rows = [("Salaries", 400000), ("Rent", 60000), ("Software", 25000),
            ("Marketing", 90000), ("Travel", 18000)]
    if version == 15:
        rows.insert(1, ("Contractors", 75000))
    start = 4
    for i, (label, amt) in enumerate(rows):
        op["B%d" % (start + i)] = label
        op["C%d" % (start + i)] = amt
        op["D%d" % (start + i)] = "=C%d*1.10" % (start + i)
    last = start + len(rows) - 1
    op["B%d" % (last + 2)] = "Total opex"
    op["C%d" % (last + 2)] = "=SUM(C%d:C%d)" % (start, last)
    op["D%d" % (last + 2)] = "=SUM(D%d:D%d)" % (start, last)

    if version == 15:
        sn = wb.create_sheet("Sensitivity")
        sn["A1"] = "Growth sensitivity"
        sn["B3"] = "Case"
        sn["C3"] = "Growth"
        for i, (label, g) in enumerate([("Low", 0.01), ("Base", 0.05), ("High", 0.09)]):
            sn["B%d" % (4 + i)] = label
            sn["C%d" % (4 + i)] = g

    wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), "model_v%d.xlsx" % version))


build(14)
build(15)
print("written")
