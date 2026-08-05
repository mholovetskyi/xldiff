"""Annotating a copy of the revised workbook.

The reviewer works in Excel. These check that the copy carries the findings
into the grid, that the original is untouched, and that the count of what could
not be attached is reported rather than swallowed.

    python tests/test_annotate.py
"""

import os
import shutil
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "cli"))

from openpyxl import load_workbook          # noqa: E402

from annotate import annotate               # noqa: E402
from engine import compare                  # noqa: E402

A = os.path.join(ROOT, "examples", "sample_before.xlsx")
B = os.path.join(ROOT, "examples", "sample_after.xlsx")

failures = []


def check(cond, label):
    print(("  ok   " if cond else "  FAIL ") + label)
    if not cond:
        failures.append(label)


def main():
    tmp = tempfile.mkdtemp()
    out = os.path.join(tmp, "reviewed.xlsx")
    before = os.path.getmtime(B)
    original = os.path.getsize(B)

    result = compare(A, B)
    findings = result["findings"]
    marked, homeless = annotate(B, findings, out)

    print("writing the annotated copy")
    check(os.path.exists(out), "the copy is written")
    check(os.path.getmtime(B) == before and os.path.getsize(B) == original,
          "and the revised workbook it came from is untouched")

    cell_findings = [f for f in findings
                     if f.ref and f.ref[:1].isalpha() and f.ref[1:].isdigit()]
    check(marked > 0 and marked <= len(cell_findings),
          "%d cells carry a comment, from %d cell findings" % (marked, len(cell_findings)))
    # Sheet-level and name-level findings have no cell to sit on. Dropping them
    # quietly would make the copy look like a complete record of the review.
    check(homeless == len(findings) - len(cell_findings),
          "the %d findings with no cell are counted, not swallowed" % homeless)

    wb = load_workbook(out)
    print("what the copy carries")
    plug = wb["Revenue build"]["D9"]
    check(plug.comment is not None, "the plug has a comment")
    text = plug.comment.text if plug.comment else ""
    check("[HIGH]" in text and "Hardcode replaced a formula" in text,
          "naming the severity and what happened")
    check("was: =D7*D8" in text, "and showing the formula it replaced")
    check(plug.fill.fgColor.rgb.endswith("F8CFCF"),
          "a high finding fills the cell so it is visible without hovering")

    live = wb["Revenue build"]["C7"]
    check(live.value == "=C6*(1+$B$4+0.02)",
          "formulas survive the round trip unchanged")

    print("several findings on one cell")
    counts = {}
    for f in findings:
        if f.ref and f.ref[:1].isalpha() and f.ref[1:].isdigit():
            counts[(f.sheet, f.ref)] = counts.get((f.sheet, f.ref), 0) + 1
    check(all(v == 1 for v in counts.values()) or
          any(wb[s][r].comment.text.count("[") > 1
              for (s, r), v in counts.items() if v > 1),
          "a cell with two findings gets one comment carrying both")

    print("the sheet with no findings")
    untouched = [c for row in wb["Sensitivity"].iter_rows() for c in row if c.comment]
    check(not untouched, "a sheet with no cell findings is left alone")

    shutil.rmtree(tmp, ignore_errors=True)
    print("")
    if failures:
        print("%d check(s) failed" % len(failures))
        return 1
    print("all checks passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
