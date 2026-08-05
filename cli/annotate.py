"""Write the findings back into a copy of the revised workbook.

The reviewer lives in the grid, not in a browser tab. A report they have to
read alongside the model is a report they read once; a comment attached to the
cell is one they meet while doing the work.

    python cli/cli.py old.xlsx new.xlsx --annotate reviewed.xlsx

The original is never touched. Everything is written to the copy.

One caveat worth knowing: openpyxl does not preserve the values Excel cached
for each formula, so the copy carries the formulas without their last results
until Excel recalculates on open. The annotated file is for reading comments
against, not for feeding back into xldiff -- comparing it would trip the
stale-cache warning, which is the tool telling the truth about what it was
handed.
"""

import re

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

_CELLREF = re.compile(r'^[A-Z]{1,3}[1-9][0-9]{0,6}$')

# Excel renders these behind the cell contents, so they have to stay light
# enough to read black text through.
FILL = {
    "high": PatternFill("solid", fgColor="F8CFCF"),
    "medium": PatternFill("solid", fgColor="FBE7C6"),
    "low": PatternFill("solid", fgColor="DCE9F7"),
}

SEV_ORDER = {"high": 0, "medium": 1, "low": 2}


def comment_text(findings):
    """One comment per cell, however many findings landed on it."""
    lines = []
    for f in findings:
        lines.append("[%s] %s" % (f.severity.upper(), f.summary))
        if f.detail:
            lines.append(f.detail)
        if f.before or f.after:
            lines.append("was: %s" % (f.before or "(empty)"))
            lines.append("now: %s" % (f.after or "(empty)"))
        if f.val_before != f.val_after and (f.val_before or f.val_after):
            lines.append("value %s -> %s" % (f.val_before or "(empty)",
                                             f.val_after or "(empty)"))
        if len(f.chain) > 1:
            lines.append("moves " + " -> ".join(f.chain[1:]))
        for o in f.outputs[:2]:
            lines.append("%s %s: %s -> %s (%.1f%%)"
                         % (o["ref"], o["label"] or "output",
                            o["before"], o["after"], o["move"] * 100))
        if f.waived:
            lines.append("waived: %s" % (f.waiver.get("reason") or "no reason given"))
        lines.append("")
    return "\n".join(lines).rstrip()


def annotate(path_new, findings, out_path):
    """Copy the revised workbook with a comment and a fill on each finding.

    Returns how many cells were annotated and how many findings had nowhere to
    go -- sheet-level and name-level findings have no cell to attach to, and
    silently dropping them would overstate the coverage of the copy.
    """
    wb = load_workbook(path_new)

    by_cell = {}
    homeless = 0
    for f in findings:
        if not f.ref or not _CELLREF.match(f.ref) or f.sheet not in wb.sheetnames:
            homeless += 1
            continue
        by_cell.setdefault((f.sheet, f.ref), []).append(f)

    for (sheet, ref), group in sorted(by_cell.items()):
        group.sort(key=lambda f: SEV_ORDER[f.severity])
        cell = wb[sheet][ref]
        text = comment_text(group)
        # Excel sizes comment boxes in points and will not grow them to fit.
        comment = Comment(text, "xldiff")
        comment.width = 340
        comment.height = min(400, 26 + 13 * text.count("\n"))
        cell.comment = comment
        worst = group[0].severity
        if not all(f.waived for f in group):
            cell.fill = FILL[worst]

    wb.save(out_path)
    return len(by_cell), homeless
