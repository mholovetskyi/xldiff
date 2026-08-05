"""Renders the comparison as one self-contained HTML file. No external assets."""

import html
import json
import os

from openpyxl.utils import get_column_letter

from engine import fmt

CONTEXT = 2
MAX_COLS = 12
SEV_COLOR = {"high": "#A32D2D", "medium": "#854F0B", "low": "#185FA5"}


def cell_pair(left, right, lr, rr, c):
    lc = left.get(lr, c) if lr else None
    rc = right.get(rr, c) if rr else None
    return {
        "c": get_column_letter(c),
        "lf": "" if lc is None else fmt(lc.formula),
        "lv": "" if lc is None else fmt(lc.value),
        "rf": "" if rc is None else fmt(rc.formula),
        "rv": "" if rc is None else fmt(rc.value),
    }


def build_hunks(name, left, right, align, sev_by_ref):
    info = align.get(name)
    if not info:
        return []
    pairs = info["pairs"]
    added, deleted = set(info["added"]), set(info["deleted"])

    rows = []
    for lr, rr in pairs:
        rows.append((lr, rr))
    for r in sorted(added):
        rows.append((None, r))
    for r in sorted(deleted):
        rows.append((r, None))
    rows.sort(key=lambda t: (t[1] if t[1] else (t[0] or 0), t[0] or 0))

    interesting = set()
    for i, (lr, rr) in enumerate(rows):
        hot = (lr is None or rr is None)
        if not hot and rr:
            for c in range(1, MAX_COLS + 1):
                if ("%s%d" % (get_column_letter(c), rr)) in sev_by_ref.get(name, {}):
                    hot = True
                    break
        if hot:
            for j in range(max(0, i - CONTEXT), min(len(rows), i + CONTEXT + 1)):
                interesting.add(j)

    ncol = min(max(left.max_col, right.max_col), MAX_COLS)
    out, prev = [], None
    for i in sorted(interesting):
        lr, rr = rows[i]
        if prev is not None and i != prev + 1:
            out.append({"gap": True})
        prev = i
        cells = [cell_pair(left, right, lr, rr, c) for c in range(1, ncol + 1)]
        for cl in cells:
            ref = "%s%d" % (cl["c"], rr) if rr else ""
            cl["sev"] = sev_by_ref.get(name, {}).get(ref, "")
        out.append({"l": lr, "r": rr, "cells": cells,
                    "state": "added" if lr is None else ("deleted" if rr is None else "")})
    return out


def render(result, out_path):
    left, right = result["left"], result["right"]
    findings = result["findings"]

    sev_by_ref = {}
    for f in findings:
        if f.ref and f.kind != "impact_capped":
            sev_by_ref.setdefault(f.sheet, {})[f.ref] = f.severity

    sheets = []
    for name in right:
        if name in left:
            hunks = build_hunks(name, left[name], right[name], result["align"], sev_by_ref)
        else:
            hunks = []
        sheets.append({
            "name": name,
            "state": "" if name in left else "added",
            "hunks": hunks,
            "n": sum(1 for f in findings if f.sheet == name),
        })
    for name in left:
        if name not in right:
            sheets.append({"name": name, "state": "deleted", "hunks": [],
                           "n": sum(1 for f in findings if f.sheet == name)})

    flist = [{
        "sev": f.severity, "kind": f.kind, "sheet": f.sheet, "ref": f.ref,
        "summary": f.summary, "detail": f.detail,
        "before": f.before, "after": f.after,
        "vb": f.val_before, "va": f.val_after,
    } for f in findings]

    counts = {s: sum(1 for f in findings if f.severity == s)
              for s in ("high", "medium", "low")}

    data = {
        "sheets": sheets, "findings": flist, "counts": counts,
        "stale": result["stale"],
        "fileA": os.path.basename(result["files"][0]),
        "fileB": os.path.basename(result["files"][1]),
    }

    with open(out_path, "w") as fh:
        fh.write(TEMPLATE.replace("__DATA__", json.dumps(data)))
    return out_path


TEMPLATE = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>xldiff report</title>
<style>
:root{--bg:#fbfaf8;--card:#fff;--line:#e4e1da;--txt:#26241f;--dim:#6d6a63;
--hi:#fcebeb;--hib:#A32D2D;--me:#faeeda;--meb:#854F0B;--lo:#e6f1fb;--lob:#185FA5;}
*{box-sizing:border-box}
body{margin:0;background:var(--bg);color:var(--txt);
font:14px/1.6 -apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif}
.wrap{max-width:1200px;margin:0 auto;padding:24px}
h1{font-size:20px;font-weight:500;margin:0 0 4px}
.sub{color:var(--dim);font-size:13px;margin-bottom:20px}
.card{background:var(--card);border:1px solid var(--line);border-radius:10px;overflow:hidden}
.bar{display:flex;gap:6px;align-items:center;padding:10px 12px;border-bottom:1px solid var(--line);flex-wrap:wrap}
.tab{padding:4px 10px;border-radius:6px;font-size:13px;cursor:pointer;border:1px solid transparent}
.tab.on{background:var(--bg);border-color:var(--line)}
.tab .n{color:var(--dim);font-size:11px;margin-left:5px}
.tab.added{background:#eaf3de;color:#3B6D11}
.tab.deleted{background:var(--hi);color:var(--hib)}
.seg{margin-left:auto;display:flex;gap:4px}
.seg span{padding:4px 10px;font-size:12px;border:1px solid var(--line);border-radius:6px;cursor:pointer}
.seg span.on{background:var(--txt);color:#fff;border-color:var(--txt)}
.panes{display:grid;grid-template-columns:1fr 1fr}
.pane{overflow:auto;max-height:520px;border-right:1px solid var(--line)}
.pane:last-child{border-right:none}
.ph{position:sticky;top:0;background:var(--card);border-bottom:1px solid var(--line);
padding:6px 10px;font-size:11px;color:var(--dim);z-index:2}
table{border-collapse:collapse;width:100%;font:11px/1.9 ui-monospace,SFMono-Regular,Menlo,monospace}
td{padding:0 6px;white-space:nowrap;border-bottom:1px solid #f2f0eb;max-width:150px;
overflow:hidden;text-overflow:ellipsis}
td.rn{color:#a8a49b;text-align:right;width:34px;background:#faf9f6;position:sticky;left:0}
tr.gap td{background:#f4f2ed;color:#a8a49b;text-align:center;font-size:10px;padding:1px}
tr.added td{background:#f2f8e9}
tr.deleted td{background:#fdf1f1}
td.high{background:var(--hi);color:var(--hib)}
td.medium{background:var(--me);color:var(--meb)}
td.low{background:var(--lo);color:var(--lob)}
tr.focus td{outline:1px solid #26241f;outline-offset:-1px}
.list{margin-top:20px}
.f{display:flex;gap:10px;padding:11px 12px;border-bottom:1px solid var(--line);cursor:pointer;align-items:flex-start}
.f:hover{background:#f4f2ed}
.dot{width:7px;height:7px;border-radius:50%;margin-top:7px;flex:none}
.f .t{font-size:13px}
.f .d{color:var(--dim);font-size:12px}
.f .code{font:11px ui-monospace,Menlo,monospace;color:var(--dim);margin-top:3px;white-space:pre-wrap}
.ref{font:11px ui-monospace,Menlo,monospace;color:var(--dim)}
.pill{display:inline-block;padding:1px 6px;border-radius:4px;font-size:11px;margin-left:6px}
.warn{background:var(--me);color:var(--meb);padding:8px 12px;font-size:12px;border-radius:8px;margin-bottom:16px}
.foot{color:var(--dim);font-size:12px;margin-top:18px}
</style></head><body><div class="wrap">
<h1>Model comparison</h1>
<div class="sub" id="sub"></div>
<div id="warn"></div>
<div class="card">
<div class="bar" id="tabs"></div>
<div class="bar" style="border-bottom:1px solid var(--line)">
<span style="font-size:12px;color:var(--dim)">View</span>
<div class="seg" style="margin-left:8px">
<span data-layer="values">Values</span><span data-layer="formulas" class="on">Formulas</span>
</div>
<span style="margin-left:auto;font-size:12px;color:var(--dim)">nothing left this machine</span>
</div>
<div class="panes">
<div class="pane" id="paneL"><div class="ph" id="hl"></div><table id="tl"></table></div>
<div class="pane" id="paneR"><div class="ph" id="hr"></div><table id="tr_"></table></div>
</div>
</div>
<div class="card list"><div id="findings"></div></div>
<div class="foot">Generated by xldiff. Findings are ordered by severity, then position.</div>
</div>
<script>
var D=__DATA__;
var layer="formulas",active=0;
var SEVC={high:"#A32D2D",medium:"#854F0B",low:"#185FA5"};

document.getElementById("sub").textContent=
 D.fileA+"  vs  "+D.fileB+"   ·   "+D.counts.high+" high, "+D.counts.medium+" medium, "+D.counts.low+" low";
if(D.stale){document.getElementById("warn").innerHTML=
 '<div class="warn">One of these files has no cached calculation values. Value impact below is incomplete. Open it in Excel, recalculate, and save.</div>';}

function tabs(){var h="";D.sheets.forEach(function(s,i){
 h+='<span class="tab '+(i==active?"on ":"")+s.state+'" data-i="'+i+'">'+esc(s.name)+
    '<span class="n">'+s.n+'</span></span>';});
 var e=document.getElementById("tabs");e.innerHTML=h;
 e.querySelectorAll(".tab").forEach(function(t){t.onclick=function(){active=+t.dataset.i;tabs();grid();};});}

function esc(s){return String(s).replace(/[&<>]/g,function(c){return{"&":"&amp;","<":"&lt;",">":"&gt;"}[c];});}

function grid(){var s=D.sheets[active];
 document.getElementById("hl").textContent=D.fileA+" · "+s.name;
 document.getElementById("hr").textContent=D.fileB+" · "+s.name;
 var L="",R="";
 if(!s.hunks.length){L=R='<tr><td style="color:#a8a49b;padding:10px">no aligned changes on this sheet</td></tr>';}
 s.hunks.forEach(function(h,i){
  if(h.gap){L+='<tr class="gap"><td colspan="20">···</td></tr>';R+='<tr class="gap"><td colspan="20">···</td></tr>';return;}
  var cls=h.state?' class="'+h.state+'"':"";
  L+='<tr'+cls+' data-h="'+i+'"><td class="rn">'+(h.l||"")+'</td>';
  R+='<tr'+cls+' data-h="'+i+'"><td class="rn">'+(h.r||"")+'</td>';
  h.cells.forEach(function(c){
   L+='<td class="'+(c.sev||"")+'">'+esc(layer=="formulas"?c.lf:c.lv)+'</td>';
   R+='<td class="'+(c.sev||"")+'">'+esc(layer=="formulas"?c.rf:c.rv)+'</td>';});
  L+='</tr>';R+='</tr>';});
 document.getElementById("tl").innerHTML=L;document.getElementById("tr_").innerHTML=R;}

function list(){var h="";
 D.findings.forEach(function(f,i){
  var code="";
  if(f.before||f.after)code=esc(f.before)+"   →   "+esc(f.after);
  var val=(f.vb!==f.va&&(f.vb||f.va))?'<span class="ref">value '+esc(f.vb)+" → "+esc(f.va)+"</span>":"";
  h+='<div class="f" data-i="'+i+'"><span class="dot" style="background:'+SEVC[f.sev]+'"></span><div>'+
     '<div class="t">'+esc(f.summary)+(f.ref?' <span class="ref">'+esc(f.sheet)+"!"+esc(f.ref)+"</span>":
     (f.sheet?' <span class="ref">'+esc(f.sheet)+"</span>":""))+"</div>"+
     (f.detail?'<div class="d">'+esc(f.detail)+"</div>":"")+
     (code?'<div class="code">'+code+"</div>":"")+
     (val?'<div class="d">'+val+"</div>":"")+"</div></div>";});
 var e=document.getElementById("findings");e.innerHTML=h;
 e.querySelectorAll(".f").forEach(function(el){el.onclick=function(){jump(D.findings[+el.dataset.i]);};});}

function jump(f){
 var idx=-1;D.sheets.forEach(function(s,i){if(s.name==f.sheet)idx=i;});
 if(idx<0)return;
 if(idx!=active){active=idx;tabs();grid();}
 if(!f.ref)return;
 var row=parseInt(f.ref.replace(/[^0-9]/g,""),10);
 var s=D.sheets[active],hi=-1;
 s.hunks.forEach(function(h,i){if(!h.gap&&h.r==row)hi=i;});
 if(hi<0)return;
 ["tl","tr_"].forEach(function(id){
  var t=document.getElementById(id);
  t.querySelectorAll("tr").forEach(function(r){r.classList.remove("focus");});
  var tgt=t.querySelector('tr[data-h="'+hi+'"]');
  if(tgt){tgt.classList.add("focus");tgt.scrollIntoView({block:"center"});}});}

document.querySelectorAll(".seg span").forEach(function(b){b.onclick=function(){
 layer=b.dataset.layer;
 document.querySelectorAll(".seg span").forEach(function(x){x.classList.remove("on");});
 b.classList.add("on");grid();};});

var pl=document.getElementById("paneL"),pr=document.getElementById("paneR"),lock=false;
function sync(a,b){a.addEventListener("scroll",function(){if(lock)return;lock=true;
 b.scrollTop=a.scrollTop;b.scrollLeft=a.scrollLeft;lock=false;});}
sync(pl,pr);sync(pr,pl);

tabs();grid();list();
</script></body></html>"""
