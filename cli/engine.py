"""xldiff engine: structural diff of two xlsx workbooks.

Reports formula changes, hardcodes, broken formula runs, and downstream
value impact. Everything runs locally. No network calls anywhere.
"""

import re
from dataclasses import dataclass, field
from difflib import SequenceMatcher

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MAX_COL = 60
MAX_ROW = 5000

_QUOTED = re.compile(r'"[^"]*"')
_REF = re.compile(r'(?<![A-Za-z0-9_$.])(\$?)([A-Z]{1,3})(\$?)([1-9][0-9]{0,6})(?![0-9(])')

IMPACT_CAP = 25
BORING_LITERALS = {0, 1, -1, 2, 12, 100, 365, 360, 1000, 0.5}


def col_to_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


_EXTLINK = re.compile(r'\[[^\[\]]+\]')


def external_links(formula):
    """Workbook references embedded in a formula, as [book.xlsx] tokens."""
    if not isinstance(formula, str):
        return []
    return sorted(set(_EXTLINK.findall(formula)))


def to_r1c1(formula, row, col, names=()):
    """Normalise an A1 formula to R1C1 so a copied formula compares equal.

    Defined names are left alone. A name like TAX1 is shaped exactly like a
    cell reference, and rewriting it relative to its own position makes the
    same formula look different once it moves down a row -- a false change
    reported on every formula that uses a name, in any file where a row was
    inserted above it.
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula
    placeholders = []

    def stash(m):
        placeholders.append(m.group(0))
        return "\x00%d\x00" % (len(placeholders) - 1)

    body = _QUOTED.sub(stash, formula)

    def repl(m):
        cabs, letters, rabs, digits = m.groups()
        if (letters + digits).upper() in names:
            return m.group(0)
        c = col_to_num(letters)
        r = int(digits)
        rpart = "R%d" % r if rabs else ("R[%d]" % (r - row) if r != row else "R")
        cpart = "C%d" % c if cabs else ("C[%d]" % (c - col) if c != col else "C")
        return rpart + cpart

    body = _REF.sub(repl, body)
    for i, p in enumerate(placeholders):
        body = body.replace("\x00%d\x00" % i, p)
    return body


def literals_in(formula):
    """Numeric constants embedded in a formula, ignoring cell refs."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return []
    body = _QUOTED.sub("", formula)
    body = _REF.sub("", body)
    out = []
    for tok in re.findall(r'(?<![A-Za-z0-9_.\]])\d+(?:\.\d+)?', body):
        try:
            v = float(tok)
        except ValueError:
            continue
        if v not in BORING_LITERALS:
            out.append(v)
    return out


@dataclass
class Cell:
    formula: object = None
    r1c1: object = None
    value: object = None

    @property
    def is_formula(self):
        return isinstance(self.formula, str) and self.formula.startswith("=")

    @property
    def empty(self):
        return self.formula is None and self.value is None


@dataclass
class Finding:
    severity: str
    kind: str
    sheet: str
    ref: str
    summary: str
    detail: str = ""
    before: str = ""
    after: str = ""
    val_before: str = ""
    val_after: str = ""
    magnitude: float = 0.0
    row_left: int = 0
    row_right: int = 0

    SEV_ORDER = {"high": 0, "medium": 1, "low": 2}

    def sort_key(self):
        return (self.SEV_ORDER[self.severity], self.sheet, self.row_right or self.row_left)


@dataclass
class SheetGrid:
    name: str
    cells: dict = field(default_factory=dict)   # (row, col) -> Cell
    max_row: int = 0
    max_col: int = 0

    def get(self, row, col):
        return self.cells.get((row, col), Cell())

    def row_cells(self, row):
        return [(c, self.cells[(row, c)]) for c in range(1, self.max_col + 1)
                if (row, c) in self.cells]


def read_names(wb):
    """Workbook-level defined names, as name -> the reference it points at.

    Sheet-local names are out of scope: they cannot be referenced from
    elsewhere, so a change to one is already visible as a formula change.
    """
    out = {}
    for name, dn in getattr(wb, "defined_names", {}).items():
        try:
            out[str(name)] = str(dn.value)
        except (AttributeError, TypeError):
            continue
    return out


def read_workbook(path):
    """Two passes: formulas, then cached values."""
    wbf = load_workbook(path, data_only=False, read_only=True)
    wbv = load_workbook(path, data_only=True, read_only=True)
    grids = {}
    stale = False
    names = read_names(wbf)
    upper = set(n.upper() for n in names)
    for name in wbf.sheetnames:
        sf, sv = wbf[name], wbv[name]
        g = SheetGrid(name)
        nrow = min(sf.max_row or 0, MAX_ROW)
        ncol = min(sf.max_column or 0, MAX_COL)
        g.max_row, g.max_col = nrow, ncol
        vals = {}
        for r in sv.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
            for c in r:
                if c.value is None or not hasattr(c, "row"):
                    continue
                vals[(c.row, c.column)] = c.value
        for r in sf.iter_rows(min_row=1, max_row=nrow, max_col=ncol):
            for c in r:
                if not hasattr(c, "row"):
                    continue
                v = vals.get((c.row, c.column))
                if c.value is None and v is None:
                    continue
                cell = Cell(formula=c.value, value=v)
                if cell.is_formula:
                    cell.r1c1 = to_r1c1(c.value, c.row, c.column, upper)
                    if v is None:
                        stale = True
                else:
                    cell.r1c1 = c.value
                g.cells[(c.row, c.column)] = cell
        grids[name] = g
    wbf.close()
    wbv.close()
    return grids, stale, names


def _collapse(shape):
    """Run-length normalise a shape string: "tnnnn" and "tnnnnn" both read "tn".

    The fingerprint records the pattern of cell kinds along the row, not how
    many columns each run happens to span. Without this an inserted column
    changes the shape of every row at once and row alignment collapses, so the
    two axes would each need the other to have been solved first.
    """
    out = []
    for ch in shape:
        if not out or out[-1] != ch:
            out.append(ch)
    return "".join(out)


def row_signature(grid, row):
    """Coarse fingerprint for alignment: label text plus a shape string.

    Deliberately ignores formula content so a row whose formula was edited
    still aligns with its counterpart.
    """
    label = ""
    shape = []
    for c in range(1, grid.max_col + 1):
        cell = grid.cells.get((row, c))
        if cell is None or cell.empty:
            shape.append(".")
            continue
        if cell.is_formula:
            shape.append("f")
        elif isinstance(cell.formula, str):
            shape.append("t")
            if not label:
                label = cell.formula.strip()[:40]
        else:
            shape.append("n")
    return "%s|%s" % (label.lower(), _collapse("".join(shape)))


def _align(lsig, rsig):
    """LCS over two fingerprint sequences, as 1-based pairs plus the strays.

    A run of mismatches is paired up positionally rather than reported as a
    wholesale delete and insert, so a block that was edited in place still
    lines up with its counterpart.
    """
    sm = SequenceMatcher(None, lsig, rsig, autojunk=False)
    pairs, added, deleted = [], [], []
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            for k in range(i2 - i1):
                pairs.append((i1 + k + 1, j1 + k + 1))
        elif tag == "replace":
            n = min(i2 - i1, j2 - j1)
            for k in range(n):
                pairs.append((i1 + k + 1, j1 + k + 1))
            for k in range(n, i2 - i1):
                deleted.append(i1 + k + 1)
            for k in range(n, j2 - j1):
                added.append(j1 + k + 1)
        elif tag == "delete":
            deleted.extend(range(i1 + 1, i2 + 1))
        elif tag == "insert":
            added.extend(range(j1 + 1, j2 + 1))
    return pairs, added, deleted, sm.ratio()


def align_rows(left, right):
    """Sequence-align rows so an inserted row does not shift the whole diff."""
    lsig = [row_signature(left, r) for r in range(1, left.max_row + 1)]
    rsig = [row_signature(right, r) for r in range(1, right.max_row + 1)]
    return _align(lsig, rsig)


def column_signature(grid, col, rows):
    """Fingerprint for a column: its header text plus a shape string.

    Read over the rows that already aligned, not over every row in the sheet.
    A single inserted row would otherwise perturb every column signature at
    once, and two sequences that share nothing align by position -- which is
    exactly the over-reporting column alignment exists to prevent.
    """
    label = ""
    shape = []
    for r in rows:
        cell = grid.cells.get((r, col))
        if cell is None or cell.empty:
            shape.append(".")
            continue
        if cell.is_formula:
            shape.append("f")
        elif isinstance(cell.formula, str):
            shape.append("t")
            if not label:
                label = cell.formula.strip()[:40]
        else:
            shape.append("n")
    return "%s|%s" % (label.lower(), "".join(shape))


def align_columns(left, right, row_pairs):
    """Sequence-align columns so an inserted column does not shift the diff."""
    lrows = [lr for lr, _ in row_pairs]
    rrows = [rr for _, rr in row_pairs]
    lsig = [column_signature(left, c, lrows) for c in range(1, left.max_col + 1)]
    rsig = [column_signature(right, c, rrows) for c in range(1, right.max_col + 1)]
    return _align(lsig, rsig)


def _walk(grid, row, col, dr, dc, limit=4):
    """Formulas adjacent to a cell in one direction, stopping at the first gap.

    A run is contiguous. Reading past a blank row or a label into the next
    block of the sheet is how unrelated formulas end up being treated as
    neighbours, which reads either as noise or as a break that is not there.
    """
    out = []
    for k in range(1, limit + 1):
        n = grid.get(row + dr * k, col + dc * k)
        if n.empty or not n.is_formula:
            break
        out.append(n.r1c1)
    return out


def find_run_break(grid, row, col, axis="row"):
    """Is this cell the odd one out in an otherwise uniform run?

    axis="row" reads the horizontal neighbours, axis="column" the vertical
    ones. Both matter: drag-filling across a row and dragging down a column
    break a deliberate exception in exactly the same way.
    """
    cell = grid.get(row, col)
    dr, dc = (0, 1) if axis == "row" else (1, 0)
    fb = _walk(grid, row, col, -dr, -dc)
    fa = _walk(grid, row, col, dr, dc)
    seen = fb + fa
    if len(seen) < 2 or len(set(seen)) != 1:
        return None

    # A column of identical formulas nearly always ends in a total that is
    # meant to differ, so a vertical break only counts when the run continues
    # on both sides of the odd cell. Rows do not need this: a row of quarters
    # is a homogeneous series to its last column, and requiring both sides
    # there would blind the tool to a break in the final period.
    if axis == "column" and (not fb or not fa):
        return None

    expected = seen[0]
    if cell.is_formula and cell.r1c1 == expected:
        return None
    return expected


def run_context(grid, row, col):
    """What this cell's neighbours agree on, and which axis they lie along.

    Horizontal is tried first: it is the more common accident, and settling on
    one axis keeps a single cell from producing two findings that describe the
    same mistake.
    """
    for axis in ("row", "column"):
        expected = find_run_break(grid, row, col, axis)
        if expected is not None:
            return expected, axis
    return None, None


_ERROR = re.compile(r'#(REF!|DIV/0!|VALUE!|NAME\?|N/A|NULL!|NUM!|SPILL!|CALC!)')


def error_text(value):
    """The Excel error a cached value carries, or None if it holds a result."""
    if isinstance(value, str):
        m = _ERROR.fullmatch(value.strip())
        if m:
            return m.group(0)
    return None


def broken_ref(formula):
    """A formula whose text has lost a reference reads #REF! literally."""
    return isinstance(formula, str) and "#REF!" in formula


def error_finding(name, ref, lc, rc, base):
    """Errors outrank every other reading of a cell.

    A formula that changed and a formula that now returns #DIV/0! are the same
    edit, but only one of them is worth waking someone for, so the error is
    what gets reported and the formula change rides along in before/after.
    """
    lb, rb = broken_ref(lc.formula), broken_ref(rc.formula)
    if rb and not lb:
        return Finding(
            "high", "broken_reference", name, ref,
            "Formula lost a reference and now reads #REF!",
            "Whatever this pointed at was deleted. The cell cannot recalculate.",
            before=fmt(lc.formula), after=fmt(rc.formula), **base)

    le, re_ = error_text(lc.value), error_text(rc.value)
    if re_ and not le:
        return Finding(
            "high", "error_introduced", name, ref,
            "This cell now evaluates to %s" % re_,
            "It held a value in the baseline.",
            before=fmt(lc.formula), after=fmt(rc.formula), **base)
    if le and not re_:
        return Finding(
            "low", "error_cleared", name, ref,
            "An error here was fixed",
            "The baseline returned %s." % le,
            before=fmt(lc.formula), after=fmt(rc.formula), **base)
    if lb and not rb:
        return Finding(
            "low", "error_cleared", name, ref,
            "A broken reference here was repaired",
            before=fmt(lc.formula), after=fmt(rc.formula), **base)
    return None


def value_delta(lc, rc):
    a, b = lc.value, rc.value
    if isinstance(a, (int, float)) and isinstance(b, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        if a == b:
            return 0.0
        return abs(b - a) / max(abs(a), 1e-9)
    return 0.0 if a == b else 1.0


def diff_sheet(name, left, right, findings, align_out=None):
    pairs, added, deleted, ratio = align_rows(left, right)
    col_pairs, col_added, col_deleted, col_ratio = align_columns(left, right, pairs)
    if align_out is not None:
        align_out[name] = {"pairs": pairs, "added": added,
                           "deleted": deleted, "ratio": ratio,
                           "col_pairs": col_pairs, "col_added": col_added,
                           "col_deleted": col_deleted, "col_ratio": col_ratio}
    if ratio < 0.35:
        findings.append(Finding(
            "low", "alignment", name, "",
            "Row alignment confidence is low (%.0f%%)" % (ratio * 100),
            "The two sheets may not be versions of each other. Findings below may be noisy."))

    for r in added:
        if any(not left.get(0, 0) for _ in []):
            pass
        label = row_signature(right, r).split("|")[0]
        findings.append(Finding(
            "low", "row_added", name, "row %d" % r,
            "Row inserted%s" % (": %s" % label if label else ""),
            row_right=r))
    for r in deleted:
        label = row_signature(left, r).split("|")[0]
        findings.append(Finding(
            "low", "row_deleted", name, "row %d" % r,
            "Row deleted%s" % (": %s" % label if label else ""),
            row_left=r))

    rrows = [rr for _, rr in pairs]
    lrows = [lr for lr, _ in pairs]
    for c in col_added:
        label = column_signature(right, c, rrows).split("|")[0]
        findings.append(Finding(
            "low", "column_added", name, "column %s" % get_column_letter(c),
            "Column inserted%s" % (": %s" % label if label else "")))
    for c in col_deleted:
        label = column_signature(left, c, lrows).split("|")[0]
        findings.append(Finding(
            "medium", "column_deleted", name, "column %s" % get_column_letter(c),
            "Column deleted%s" % (": %s" % label if label else "")))

    for lr, rr in pairs:
        for lcol, rcol in col_pairs:
            lc, rc = left.get(lr, lcol), right.get(rr, rcol)
            if lc.empty and rc.empty:
                continue
            c = rcol
            ref = "%s%d" % (get_column_letter(rcol), rr)
            same_logic = lc.r1c1 == rc.r1c1
            vb, va = fmt(lc.value), fmt(rc.value)
            mag = value_delta(lc, rc)

            err = error_finding(name, ref, lc, rc, dict(
                val_before=vb, val_after=va, magnitude=mag,
                row_left=lr, row_right=rr))
            if err is not None:
                findings.append(err)
                continue

            # A formula that now reads from a different workbook is a change of
            # source, not of logic, and the diff of the text buries it.
            lx, rx = external_links(lc.formula), external_links(rc.formula)
            if lx != rx and (lx or rx):
                findings.append(Finding(
                    "high", "external_link_changed", name, ref,
                    "Formula reads from a different workbook",
                    "Was %s, now %s." % (", ".join(lx) or "no external link",
                                         ", ".join(rx) or "no external link"),
                    before=fmt(lc.formula), after=fmt(rc.formula),
                    val_before=vb, val_after=va, magnitude=mag,
                    row_left=lr, row_right=rr))
                continue

            if same_logic:
                if lc.value != rc.value and not lc.empty:
                    findings.append(Finding(
                        "medium", "impact", name, ref,
                        "Value moved with no edit to this cell",
                        "Downstream of a change elsewhere.",
                        before="", after="",
                        val_before=vb, val_after=va, magnitude=mag,
                        row_left=lr, row_right=rr))
                continue

            if lc.empty:
                findings.append(Finding(
                    "low", "cell_added", name, ref, "Cell added",
                    after=fmt(rc.formula),
                    val_before=vb, val_after=va, magnitude=mag,
                    row_left=lr, row_right=rr))
                continue
            if rc.empty:
                findings.append(Finding(
                    "medium", "cell_deleted", name, ref, "Cell cleared",
                    before=fmt(lc.formula),
                    val_before=vb, val_after=va, magnitude=mag,
                    row_left=lr, row_right=rr))
                continue

            if lc.is_formula and not rc.is_formula:
                expected, axis = run_context(right, rr, rcol)
                findings.append(Finding(
                    "high", "hardcode", name, ref,
                    "Hardcode replaced a formula",
                    ("The %s it sits in is still calculated." % axis) if expected else
                    "This cell no longer recalculates.",
                    before=fmt(lc.formula), after=fmt(rc.formula),
                    val_before=vb, val_after=va, magnitude=mag,
                    row_left=lr, row_right=rr))
                continue

            if not lc.is_formula and rc.is_formula:
                findings.append(Finding(
                    "medium", "constant_to_formula", name, ref,
                    "Constant replaced by a formula",
                    before=fmt(lc.formula), after=fmt(rc.formula),
                    val_before=vb, val_after=va, magnitude=mag,
                    row_left=lr, row_right=rr))
                continue

            if lc.is_formula and rc.is_formula:
                expected, axis = run_context(right, rr, rcol)
                was_break, was_axis = run_context(left, lr, lcol)
                if expected is not None:
                    findings.append(Finding(
                        "high", "run_break", name, ref,
                        "Formula breaks the run in its %s" % axis,
                        "Neighbours use %s" % expected,
                        before=fmt(lc.formula), after=fmt(rc.formula),
                        val_before=vb, val_after=va, magnitude=mag,
                        row_left=lr, row_right=rr))
                elif was_break is not None:
                    findings.append(Finding(
                        "low", "run_repaired", name, ref,
                        "Formula now matches the rest of its %s" % was_axis,
                        before=fmt(lc.formula), after=fmt(rc.formula),
                        val_before=vb, val_after=va, magnitude=mag,
                        row_left=lr, row_right=rr))
                else:
                    new_lits = set(literals_in(rc.formula)) - set(literals_in(lc.formula))
                    sev = "high" if new_lits else "medium"
                    detail = ("New numeric literal: %s" %
                              ", ".join(fmt(x) for x in sorted(new_lits))) if new_lits else ""
                    findings.append(Finding(
                        sev, "formula_changed", name, ref,
                        "Formula changed" + (" and gained a hardcoded number" if new_lits else ""),
                        detail, before=fmt(lc.formula), after=fmt(rc.formula),
                        val_before=vb, val_after=va, magnitude=mag,
                        row_left=lr, row_right=rr))
                continue

            findings.append(Finding(
                "medium", "constant_changed", name, ref, "Constant changed",
                before=fmt(lc.formula), after=fmt(rc.formula),
                val_before=vb, val_after=va, magnitude=mag,
                row_left=lr, row_right=rr))


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v == int(v) and abs(v) < 1e15:
            return "{:,}".format(int(v))
        return "{:,.4f}".format(v).rstrip("0").rstrip(".")
    if isinstance(v, int):
        return "{:,}".format(v)
    return str(v)


def diff_names(names_a, names_b, findings):
    """Defined names, which break formulas silently when they move.

    A deleted name leaves every formula that used it unresolvable, and a
    repointed name changes what those formulas mean without changing a single
    character of their text -- the one edit a formula diff cannot see.
    """
    for n in sorted(set(names_b) - set(names_a)):
        findings.append(Finding(
            "low", "name_added", "", n, "Named range added: %s" % n,
            after=names_b[n]))
    for n in sorted(set(names_a) - set(names_b)):
        findings.append(Finding(
            "high", "name_deleted", "", n, "Named range deleted: %s" % n,
            "Any formula that referred to it can no longer resolve.",
            before=names_a[n]))
    for n in sorted(set(names_a) & set(names_b)):
        if names_a[n] != names_b[n]:
            findings.append(Finding(
                "high", "name_changed", "", n,
                "Named range %s now points somewhere else" % n,
                "Formulas using it changed meaning without changing text.",
                before=names_a[n], after=names_b[n]))


def compare(path_a, path_b):
    left, stale_a, names_a = read_workbook(path_a)
    right, stale_b, names_b = read_workbook(path_b)
    findings = []
    align = {}

    diff_names(names_a, names_b, findings)

    for name in right:
        if name not in left:
            findings.append(Finding(
                "medium", "sheet_added", name, "", "Sheet added: %s" % name))
    for name in left:
        if name not in right:
            findings.append(Finding(
                "high", "sheet_deleted", name, "", "Sheet deleted: %s" % name))
    for name in left:
        if name in right:
            diff_sheet(name, left[name], right[name], findings, align)

    impacts = [f for f in findings if f.kind == "impact"]
    if len(impacts) > IMPACT_CAP:
        impacts.sort(key=lambda f: -f.magnitude)
        keep = set(id(f) for f in impacts[:IMPACT_CAP])
        suppressed = len(impacts) - IMPACT_CAP
        findings = [f for f in findings if f.kind != "impact" or id(f) in keep]
        findings.append(Finding(
            "low", "impact_capped", "", "",
            "%d further cells moved in value with no edit" % suppressed,
            "Showing the %d largest by relative change." % IMPACT_CAP))

    findings.sort(key=lambda f: f.sort_key())
    return {
        "left": left, "right": right, "findings": findings, "align": align,
        "stale": stale_a or stale_b,
        "files": (path_a, path_b),
        "names": (names_a, names_b),
    }
